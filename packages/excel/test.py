from openpyxl import load_workbook
from openpyxl.styles import colors, Font, Fill, NamedStyle
from openpyxl.styles import PatternFill, Border, Side, Alignment


workbook = load_workbook('/Users/lokalone/Downloads/test.xlsx')
sheet_names = workbook.sheetnames
print(sheet_names)
wooksheet = workbook[sheet_names[0]];
wooksheet['A1'] = 42
wooksheet.append([1, 2, 3])
workbook.save('./5a.xlsx')